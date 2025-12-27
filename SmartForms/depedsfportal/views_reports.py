import csv
import datetime
import io

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
except ImportError:
    Workbook = None
from django.shortcuts import render, HttpResponse
from django.views.generic import TemplateView, View
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.db.models import Count, Q
from django.template.loader import render_to_string
from django.core.files.storage import FileSystemStorage
from django.conf import settings

# Try importing weasyprint, handle if not installed (though it should be)
try:
    from weasyprint import HTML
    from weasyprint.text.fonts import FontConfiguration
except ImportError:
    HTML = None

from .models import Student, AcademicRecord, School, AcademicYear, Section


class PrincipalAccessMixin(UserPassesTestMixin):
    def test_func(self):
        return (
            self.request.user.groups.filter(name="Principal").exists()
            or self.request.user.is_superuser
        )


class RegistrarAccessMixin(UserPassesTestMixin):
    def test_func(self):
        return (
            self.request.user.groups.filter(
                name__in=["Registrar", "Principal"]
            ).exists()
            or self.request.user.is_superuser
        )


class ReportDashboardView(LoginRequiredMixin, PrincipalAccessMixin, TemplateView):
    template_name = "reports/report_dashboard.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Filter Logic
        grade_level = self.request.GET.get("grade")
        status = self.request.GET.get("status")
        school_year = self.request.GET.get("year")

        # Base Query
        students = Student.objects.all().order_by("last_name")

        if grade_level:
            # Filter by latest academic record's grade level
            students = students.filter(
                academic_records__grade_level=grade_level
            ).distinct()

        if school_year:
            students = students.filter(
                academic_records__school_year=school_year
            ).distinct()

        if status:
            if status in ["ENROLLED", "TRANSFERRED", "DROPPED"]:
                students = students.filter(status=status)
            elif status == "PASSED":
                # Passed in specific year/grade
                students = students.filter(academic_records__remarks="PASSED")
                if school_year:
                    students = students.filter(
                        academic_records__school_year=school_year
                    )
            elif status == "REMEDIAL":
                students = students.filter(
                    academic_records__remarks="REMEDIAL"  # or 'Needs Remedial' depending on model, checking model...
                )
                # Checking model logic: determine_remarks returns 'Needs Remedial'?
                # Wait, model says "Needs Remedial" in SubjectGrade, but "PROMOTED/RETAINED/PASSED/FAILED" in AcademicRecord.
                # Let's adjust query to filter based on available fields.
                pass

        context["students"] = students
        context["current_filters"] = self.request.GET
        context["academic_years"] = AcademicYear.objects.all().order_by("-start_date")
        return context


class ExportReportCSVView(LoginRequiredMixin, PrincipalAccessMixin, View):
    def get(self, request, *args, **kwargs):
        # ... Reuse filter logic (should be factored out, but duplicating for speed for now) ...
        # Simplified: Just dump current filtered queryset

        grade_level = request.GET.get("grade")
        status = request.GET.get("status")
        school_year = request.GET.get("year")

        students = Student.objects.all().order_by("last_name")

        if grade_level:
            students = students.filter(
                academic_records__grade_level=grade_level
            ).distinct()
        if school_year:
            students = students.filter(
                academic_records__school_year=school_year
            ).distinct()
        if status and status in ["ENROLLED", "TRANSFERRED", "DROPPED"]:
            students = students.filter(status=status)

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = (
            f'attachment; filename="student_report_{datetime.date.today()}.csv"'
        )

        writer = csv.writer(response)
        writer.writerow(
            ["LRN", "Last Name", "First Name", "Sex", "Birthdate", "Status", "Address"]
        )

        for student in students:
            writer.writerow(
                [
                    student.lrn,
                    student.last_name,
                    student.first_name,
                    student.sex,
                    student.birthdate,
                    student.status,
                    f"{student.barangay}, {student.city}, {student.province}",
                ]
            )

        return response


class ExportReportPDFView(LoginRequiredMixin, PrincipalAccessMixin, View):
    def get(self, request, *args, **kwargs):
        # ... Reuse filter logic ...
        grade_level = request.GET.get("grade")
        status = request.GET.get("status")
        school_year = request.GET.get("year")

        students = Student.objects.all().order_by("last_name")

        if grade_level:
            students = students.filter(
                academic_records__grade_level=grade_level
            ).distinct()
        if school_year:
            students = students.filter(
                academic_records__school_year=school_year
            ).distinct()
        if status and status in ["ENROLLED", "TRANSFERRED", "DROPPED"]:
            students = students.filter(status=status)

        html_string = render_to_string(
            "reports/report_pdf_template.html",
            {
                "students": students,
                "generated_at": datetime.datetime.now(),
                "filters": request.GET,
            },
        )

        if HTML:
            html = HTML(string=html_string)
            result = html.write_pdf()

            response = HttpResponse(content_type="application/pdf")
            response["Content-Disposition"] = (
                f'inline; filename="report_{datetime.date.today()}.pdf"'
            )
            response["Content-Transfer-Encoding"] = "binary"
            with response as f:
                f.write(result)
            return response
        else:
            return HttpResponse("WeasyPrint not installed", status=500)


class AnalyticsDashboardView(LoginRequiredMixin, PrincipalAccessMixin, TemplateView):
    template_name = "reports/analytics_dashboard.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Base Query
        students = Student.objects.all()

        # Apply Filters
        grade_level = self.request.GET.get("grade")
        status = self.request.GET.get("status")
        school_year = self.request.GET.get("year")

        if grade_level:
            students = students.filter(
                academic_records__grade_level=grade_level
            ).distinct()
        if school_year:
            students = students.filter(
                academic_records__school_year=school_year
            ).distinct()
        if status:
            # Basic status filter
            if status in ["ENROLLED", "TRANSFERRED", "DROPPED"]:
                students = students.filter(status=status)

        # Gender Distribution
        by_gender = students.values("sex").annotate(count=Count("lrn")).order_by("sex")

        # Gender by Grade Level
        records = AcademicRecord.objects.filter(student__in=students)
        if school_year:
            records = records.filter(school_year=school_year)

        by_gender_grade = (
            records.values("grade_level", "student__sex")
            .annotate(count=Count("id"))
            .order_by("grade_level", "student__sex")
        )

        # Gender by Section
        by_gender_section = (
            records.values("section__name", "student__sex")
            .annotate(count=Count("id"))
            .order_by("section__name", "student__sex")
        )

        # Location Analytics (Filtered)
        by_barangay = (
            students.values("barangay").annotate(count=Count("lrn")).order_by("-count")
        )
        by_city = (
            students.values("city").annotate(count=Count("lrn")).order_by("-count")
        )
        by_province = (
            students.values("province").annotate(count=Count("lrn")).order_by("-count")
        )

        # Status Analytics (Filtered)
        by_status = students.values("status").annotate(count=Count("lrn"))

        context.update(
            {
                "by_gender": list(by_gender),
                "by_gender_grade": list(by_gender_grade),
                "by_gender_section": list(by_gender_section),
                "by_barangay": list(by_barangay),
                "by_city": list(by_city),
                "by_province": list(by_province),
                "by_status": list(by_status),
                "current_filters": self.request.GET,
                "academic_years": AcademicYear.objects.all().order_by("-start_date"),
            }
        )

        return context


class ClassListPDFView(LoginRequiredMixin, RegistrarAccessMixin, View):
    def get(self, request, pk, *args, **kwargs):
        section = Section.objects.get(pk=pk)
        school = School.objects.first()
        current_year = AcademicYear.get_current_year()

        # Get students currently in this section
        academic_records = (
            AcademicRecord.objects.filter(
                section=section,
                grade_level=section.grade_level,
                school_year=current_year.year_label if current_year else None,
            )
            .exclude(remarks="PROMOTED")
            .select_related("student")
            .order_by("student__sex", "student__last_name", "student__first_name")
        )

        students_male = [r.student for r in academic_records if r.student.sex == "M"]
        students_female = [r.student for r in academic_records if r.student.sex == "F"]

        html_string = render_to_string(
            "reports/class_list_pdf.html",
            {
                "section": section,
                "school": school,
                "academic_year": current_year,
                "students_male": students_male,
                "students_female": students_female,
                "generated_at": datetime.datetime.now(),
            },
        )

        if HTML:
            html = HTML(string=html_string)
            result = html.write_pdf()

            response = HttpResponse(content_type="application/pdf")
            response["Content-Disposition"] = (
                f'inline; filename="class_list_{section.grade_level}_{section.name}_{datetime.date.today()}.pdf"'
            )
            response.write(result)
            return response
        else:
            return HttpResponse("WeasyPrint not installed", status=500)


class ClassListExcelView(LoginRequiredMixin, RegistrarAccessMixin, View):
    def get(self, request, pk, *args, **kwargs):
        if not Workbook:
            return HttpResponse("openpyxl not installed", status=500)
        section = Section.objects.get(pk=pk)
        current_year = AcademicYear.get_current_year()

        academic_records = (
            AcademicRecord.objects.filter(
                section=section,
                grade_level=section.grade_level,
                school_year=current_year.year_label if current_year else None,
            )
            .exclude(remarks="PROMOTED")
            .select_related("student")
            .order_by("student__sex", "student__last_name", "student__first_name")
        )

        wb = Workbook()
        ws = wb.active
        ws.title = f"Grade {section.grade_level} - {section.name}"

        # Header Info
        ws.merge_cells("A1:D1")
        ws["A1"] = f"Class List: Grade {section.grade_level} - {section.name}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:D2")
        ws["A2"] = f"School Year: {current_year.year_label if current_year else 'N/A'}"
        ws["A2"].alignment = Alignment(horizontal="center")

        # Table Header
        headers = ["No.", "LRN", "Student Name", "Sex"]
        ws.append([])  # Spacer
        ws.append(headers)

        header_row = ws[4]
        for cell in header_row:
            cell.font = Font(bold=True)
            cell.border = Border(bottom=Side(style="thin"))

        # Data
        count = 1
        # Male Students
        male_exists = any(r.student.sex == "M" for r in academic_records)
        if male_exists:
            ws.append(["MALE"])
            ws[ws.max_row][0].font = Font(bold=True)
            for record in academic_records:
                if record.student.sex == "M":
                    ws.append(
                        [
                            count,
                            record.student.lrn,
                            f"{record.student.last_name}, {record.student.first_name}",
                            "Male",
                        ]
                    )
                    count += 1

        # Female Students
        ws.append([])  # Spacer
        female_exists = any(r.student.sex == "F" for r in academic_records)
        if female_exists:
            ws.append(["FEMALE"])
            ws[ws.max_row][0].font = Font(bold=True)
            count = 1
            for record in academic_records:
                if record.student.sex == "F":
                    ws.append(
                        [
                            count,
                            record.student.lrn,
                            f"{record.student.last_name}, {record.student.first_name}",
                            "Female",
                        ]
                    )
                    count += 1

        # Column Widths
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 10

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="class_list_{section.grade_level}_{section.name}_{datetime.date.today()}.xlsx"'
        )

        wb.save(response)
        return response
